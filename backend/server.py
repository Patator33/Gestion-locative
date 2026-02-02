from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, BackgroundTasks, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse, FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
from passlib.context import CryptContext
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import asyncio
import shutil
import base64

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Create uploads directory
UPLOADS_DIR = ROOT_DIR / 'uploads'
UPLOADS_DIR.mkdir(exist_ok=True)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Scheduler for automated reminders
scheduler = AsyncIOScheduler()

# JWT Configuration
SECRET_KEY = os.environ.get('JWT_SECRET', 'rent-maestro-secret-key-2024')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 24

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Security
security = HTTPBearer()

# Create the main app
app = FastAPI(title="RentMaestro API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ==================== MODELS ====================

class UserBase(BaseModel):
    email: EmailStr
    name: str

class UserCreate(UserBase):
    password: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class User(UserBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Token(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict

class PropertyBase(BaseModel):
    name: str
    address: str
    city: str
    postal_code: str
    property_type: str  # apartment, house, studio, etc.
    surface: float  # m²
    rooms: int
    rent_amount: float
    charges: float = 0
    description: Optional[str] = None
    image_url: Optional[str] = None

class PropertyCreate(PropertyBase):
    pass

class Property(PropertyBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    is_occupied: bool = False
    current_tenant_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TenantBase(BaseModel):
    first_name: str
    last_name: str
    email: EmailStr
    phone: str
    birth_date: Optional[str] = None
    profession: Optional[str] = None
    emergency_contact: Optional[str] = None
    notes: Optional[str] = None

class TenantCreate(TenantBase):
    pass

class Tenant(TenantBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    current_property_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LeaseBase(BaseModel):
    property_id: str
    tenant_id: str
    start_date: str  # ISO format
    end_date: Optional[str] = None
    rent_amount: float
    charges: float = 0
    deposit: float
    payment_day: int = 1  # Day of month for rent payment
    notes: Optional[str] = None

class LeaseCreate(LeaseBase):
    pass

class Lease(LeaseBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PaymentBase(BaseModel):
    lease_id: str
    amount: float
    payment_date: str  # ISO format
    period_month: int  # 1-12
    period_year: int
    payment_method: str = "virement"  # virement, cheque, especes, cb
    notes: Optional[str] = None

class PaymentCreate(PaymentBase):
    pass

class Payment(PaymentBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    status: str = "paid"  # paid, pending, late
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class VacancyBase(BaseModel):
    property_id: str
    start_date: str  # ISO format
    end_date: Optional[str] = None
    reason: Optional[str] = None

class VacancyCreate(VacancyBase):
    pass

class Vacancy(VacancyBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class NotificationSettingsBase(BaseModel):
    late_payment: bool = True
    late_payment_days: int = 5
    lease_ending: bool = True
    lease_ending_days: int = 60
    vacancy_alert: bool = True
    vacancy_alert_days: int = 30
    # Email settings
    email_reminders: bool = False
    reminder_frequency: str = "weekly"  # daily, weekly, monthly
    smtp_email: Optional[str] = None
    smtp_password: Optional[str] = None
    smtp_configured: bool = False

class NotificationSettingsCreate(NotificationSettingsBase):
    pass

class NotificationSettings(NotificationSettingsBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Notification(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    type: str  # late_payment, lease_ending, vacancy
    title: str
    message: str
    is_read: bool = False
    related_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Document model for file uploads
class DocumentBase(BaseModel):
    name: str
    document_type: str  # bail, etat_lieux_entree, etat_lieux_sortie, attestation, autre
    related_type: str  # property, tenant, lease
    related_id: str
    notes: Optional[str] = None

class DocumentCreate(DocumentBase):
    pass

class Document(DocumentBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    filename: str
    file_size: int
    mime_type: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Calendar event model
class CalendarEvent(BaseModel):
    id: str
    title: str
    date: str
    type: str  # payment_due, lease_end, vacancy_start
    related_id: Optional[str] = None
    property_name: Optional[str] = None
    tenant_name: Optional[str] = None
    amount: Optional[float] = None

# Team/Organization model for multi-user collaboration
class TeamBase(BaseModel):
    name: str
    description: Optional[str] = None

class TeamCreate(TeamBase):
    pass

class Team(TeamBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    owner_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Team member model
class TeamMemberBase(BaseModel):
    role: str = "member"  # owner, admin, member, viewer

class TeamMember(TeamMemberBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    team_id: str
    user_id: str
    invited_by: str
    joined_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Team invitation model
class TeamInvitationBase(BaseModel):
    email: EmailStr
    role: str = "member"

class TeamInvitation(TeamInvitationBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    team_id: str
    invited_by: str
    status: str = "pending"  # pending, accepted, declined, expired
    token: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    expires_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc) + timedelta(days=7))

# Audit log model for history tracking
class AuditLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    team_id: Optional[str] = None
    action: str  # create, update, delete
    entity_type: str  # property, tenant, lease, payment, document, vacancy
    entity_id: str
    entity_name: str
    changes: Optional[dict] = None  # For updates: {field: {old: x, new: y}}
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ==================== AUTH HELPERS ====================

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Token invalide")
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="Utilisateur non trouvé")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expiré")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalide")

# ==================== AUDIT LOG HELPER ====================

async def create_audit_log(
    user_id: str,
    user_name: str,
    action: str,
    entity_type: str,
    entity_id: str,
    entity_name: str,
    team_id: str = None,
    changes: dict = None
):
    """Create an audit log entry for tracking changes"""
    log = AuditLog(
        user_id=user_id,
        user_name=user_name,
        team_id=team_id,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        entity_name=entity_name,
        changes=changes
    )
    log_dict = log.model_dump()
    log_dict['created_at'] = log_dict['created_at'].isoformat()
    await db.audit_logs.insert_one(log_dict)
    return log

def get_changes(old_data: dict, new_data: dict, fields_to_track: list) -> dict:
    """Compare old and new data to get changes"""
    changes = {}
    for field in fields_to_track:
        old_val = old_data.get(field)
        new_val = new_data.get(field)
        if old_val != new_val:
            changes[field] = {"old": old_val, "new": new_val}
    return changes if changes else None

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register", response_model=Token)
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email déjà utilisé")
    
    user = User(email=user_data.email, name=user_data.name)
    user_dict = user.model_dump()
    user_dict['password'] = get_password_hash(user_data.password)
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    
    await db.users.insert_one(user_dict)
    
    # Create default notification settings
    notif_settings = NotificationSettings(user_id=user.id)
    notif_dict = notif_settings.model_dump()
    notif_dict['created_at'] = notif_dict['created_at'].isoformat()
    await db.notification_settings.insert_one(notif_dict)
    
    token = create_access_token({"sub": user.id})
    return Token(
        access_token=token,
        user={"id": user.id, "email": user.email, "name": user.name}
    )

@api_router.post("/auth/login", response_model=Token)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user['password']):
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    
    token = create_access_token({"sub": user['id']})
    return Token(
        access_token=token,
        user={"id": user['id'], "email": user['email'], "name": user['name']}
    )

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return current_user

# ==================== PROPERTIES ROUTES ====================

@api_router.post("/properties", response_model=dict)
async def create_property(property_data: PropertyCreate, current_user: dict = Depends(get_current_user)):
    property_obj = Property(**property_data.model_dump(), user_id=current_user['id'])
    doc = property_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.properties.insert_one(doc)
    return {"id": property_obj.id, "message": "Bien créé avec succès"}

@api_router.get("/properties", response_model=List[dict])
async def get_properties(current_user: dict = Depends(get_current_user)):
    properties = await db.properties.find({"user_id": current_user['id']}, {"_id": 0}).to_list(1000)
    return properties

@api_router.get("/properties/{property_id}", response_model=dict)
async def get_property(property_id: str, current_user: dict = Depends(get_current_user)):
    property_doc = await db.properties.find_one(
        {"id": property_id, "user_id": current_user['id']}, {"_id": 0}
    )
    if not property_doc:
        raise HTTPException(status_code=404, detail="Bien non trouvé")
    return property_doc

@api_router.put("/properties/{property_id}", response_model=dict)
async def update_property(property_id: str, property_data: PropertyCreate, current_user: dict = Depends(get_current_user)):
    result = await db.properties.update_one(
        {"id": property_id, "user_id": current_user['id']},
        {"$set": property_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Bien non trouvé")
    return {"message": "Bien mis à jour avec succès"}

@api_router.delete("/properties/{property_id}")
async def delete_property(property_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.properties.delete_one({"id": property_id, "user_id": current_user['id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bien non trouvé")
    return {"message": "Bien supprimé avec succès"}

# ==================== TENANTS ROUTES ====================

@api_router.post("/tenants", response_model=dict)
async def create_tenant(tenant_data: TenantCreate, current_user: dict = Depends(get_current_user)):
    tenant_obj = Tenant(**tenant_data.model_dump(), user_id=current_user['id'])
    doc = tenant_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.tenants.insert_one(doc)
    return {"id": tenant_obj.id, "message": "Locataire créé avec succès"}

@api_router.get("/tenants", response_model=List[dict])
async def get_tenants(current_user: dict = Depends(get_current_user)):
    tenants = await db.tenants.find({"user_id": current_user['id']}, {"_id": 0}).to_list(1000)
    return tenants

@api_router.get("/tenants/{tenant_id}", response_model=dict)
async def get_tenant(tenant_id: str, current_user: dict = Depends(get_current_user)):
    tenant_doc = await db.tenants.find_one(
        {"id": tenant_id, "user_id": current_user['id']}, {"_id": 0}
    )
    if not tenant_doc:
        raise HTTPException(status_code=404, detail="Locataire non trouvé")
    return tenant_doc

@api_router.put("/tenants/{tenant_id}", response_model=dict)
async def update_tenant(tenant_id: str, tenant_data: TenantCreate, current_user: dict = Depends(get_current_user)):
    result = await db.tenants.update_one(
        {"id": tenant_id, "user_id": current_user['id']},
        {"$set": tenant_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Locataire non trouvé")
    return {"message": "Locataire mis à jour avec succès"}

@api_router.delete("/tenants/{tenant_id}")
async def delete_tenant(tenant_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.tenants.delete_one({"id": tenant_id, "user_id": current_user['id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Locataire non trouvé")
    return {"message": "Locataire supprimé avec succès"}

# ==================== LEASES ROUTES ====================

@api_router.post("/leases", response_model=dict)
async def create_lease(lease_data: LeaseCreate, current_user: dict = Depends(get_current_user)):
    # Verify property and tenant exist
    property_doc = await db.properties.find_one({"id": lease_data.property_id, "user_id": current_user['id']})
    if not property_doc:
        raise HTTPException(status_code=404, detail="Bien non trouvé")
    
    tenant_doc = await db.tenants.find_one({"id": lease_data.tenant_id, "user_id": current_user['id']})
    if not tenant_doc:
        raise HTTPException(status_code=404, detail="Locataire non trouvé")
    
    # Create lease
    lease_obj = Lease(**lease_data.model_dump(), user_id=current_user['id'])
    doc = lease_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.leases.insert_one(doc)
    
    # Update property status
    await db.properties.update_one(
        {"id": lease_data.property_id},
        {"$set": {"is_occupied": True, "current_tenant_id": lease_data.tenant_id}}
    )
    
    # Update tenant current property
    await db.tenants.update_one(
        {"id": lease_data.tenant_id},
        {"$set": {"current_property_id": lease_data.property_id}}
    )
    
    # End any active vacancy
    await db.vacancies.update_many(
        {"property_id": lease_data.property_id, "is_active": True},
        {"$set": {"is_active": False, "end_date": lease_data.start_date}}
    )
    
    return {"id": lease_obj.id, "message": "Bail créé avec succès"}

@api_router.get("/leases", response_model=List[dict])
async def get_leases(current_user: dict = Depends(get_current_user)):
    leases = await db.leases.find({"user_id": current_user['id']}, {"_id": 0}).to_list(1000)
    # Enrich with property and tenant info
    for lease in leases:
        property_doc = await db.properties.find_one({"id": lease['property_id']}, {"_id": 0, "name": 1, "address": 1})
        tenant_doc = await db.tenants.find_one({"id": lease['tenant_id']}, {"_id": 0, "first_name": 1, "last_name": 1})
        lease['property'] = property_doc
        lease['tenant'] = tenant_doc
    return leases

@api_router.get("/leases/{lease_id}", response_model=dict)
async def get_lease(lease_id: str, current_user: dict = Depends(get_current_user)):
    lease_doc = await db.leases.find_one(
        {"id": lease_id, "user_id": current_user['id']}, {"_id": 0}
    )
    if not lease_doc:
        raise HTTPException(status_code=404, detail="Bail non trouvé")
    return lease_doc

@api_router.put("/leases/{lease_id}/terminate")
async def terminate_lease(lease_id: str, end_date: str, current_user: dict = Depends(get_current_user)):
    lease_doc = await db.leases.find_one({"id": lease_id, "user_id": current_user['id']})
    if not lease_doc:
        raise HTTPException(status_code=404, detail="Bail non trouvé")
    
    # Update lease
    await db.leases.update_one(
        {"id": lease_id},
        {"$set": {"is_active": False, "end_date": end_date}}
    )
    
    # Update property
    await db.properties.update_one(
        {"id": lease_doc['property_id']},
        {"$set": {"is_occupied": False, "current_tenant_id": None}}
    )
    
    # Update tenant
    await db.tenants.update_one(
        {"id": lease_doc['tenant_id']},
        {"$set": {"current_property_id": None}}
    )
    
    # Create vacancy
    vacancy_obj = Vacancy(
        property_id=lease_doc['property_id'],
        start_date=end_date,
        user_id=current_user['id'],
        reason="Fin de bail"
    )
    doc = vacancy_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.vacancies.insert_one(doc)
    
    return {"message": "Bail résilié avec succès"}

# ==================== PAYMENTS ROUTES ====================

@api_router.post("/payments", response_model=dict)
async def create_payment(payment_data: PaymentCreate, current_user: dict = Depends(get_current_user)):
    # Verify lease exists
    lease_doc = await db.leases.find_one({"id": payment_data.lease_id, "user_id": current_user['id']})
    if not lease_doc:
        raise HTTPException(status_code=404, detail="Bail non trouvé")
    
    payment_obj = Payment(**payment_data.model_dump(), user_id=current_user['id'])
    doc = payment_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.payments.insert_one(doc)
    return {"id": payment_obj.id, "message": "Paiement enregistré avec succès"}

@api_router.get("/payments", response_model=List[dict])
async def get_payments(current_user: dict = Depends(get_current_user)):
    payments = await db.payments.find({"user_id": current_user['id']}, {"_id": 0}).to_list(1000)
    # Enrich with lease info
    for payment in payments:
        lease_doc = await db.leases.find_one({"id": payment['lease_id']}, {"_id": 0})
        if lease_doc:
            property_doc = await db.properties.find_one({"id": lease_doc['property_id']}, {"_id": 0, "name": 1})
            tenant_doc = await db.tenants.find_one({"id": lease_doc['tenant_id']}, {"_id": 0, "first_name": 1, "last_name": 1})
            payment['property'] = property_doc
            payment['tenant'] = tenant_doc
    return payments

@api_router.get("/payments/lease/{lease_id}", response_model=List[dict])
async def get_lease_payments(lease_id: str, current_user: dict = Depends(get_current_user)):
    payments = await db.payments.find(
        {"lease_id": lease_id, "user_id": current_user['id']}, {"_id": 0}
    ).to_list(1000)
    return payments

@api_router.delete("/payments/{payment_id}")
async def delete_payment(payment_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.payments.delete_one({"id": payment_id, "user_id": current_user['id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Paiement non trouvé")
    return {"message": "Paiement supprimé avec succès"}

# ==================== VACANCIES ROUTES ====================

@api_router.post("/vacancies", response_model=dict)
async def create_vacancy(vacancy_data: VacancyCreate, current_user: dict = Depends(get_current_user)):
    vacancy_obj = Vacancy(**vacancy_data.model_dump(), user_id=current_user['id'])
    doc = vacancy_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.vacancies.insert_one(doc)
    return {"id": vacancy_obj.id, "message": "Vacance créée avec succès"}

@api_router.get("/vacancies", response_model=List[dict])
async def get_vacancies(current_user: dict = Depends(get_current_user)):
    vacancies = await db.vacancies.find({"user_id": current_user['id']}, {"_id": 0}).to_list(1000)
    # Enrich with property info
    for vacancy in vacancies:
        property_doc = await db.properties.find_one({"id": vacancy['property_id']}, {"_id": 0, "name": 1, "address": 1})
        vacancy['property'] = property_doc
    return vacancies

@api_router.put("/vacancies/{vacancy_id}/end")
async def end_vacancy(vacancy_id: str, end_date: str, current_user: dict = Depends(get_current_user)):
    result = await db.vacancies.update_one(
        {"id": vacancy_id, "user_id": current_user['id']},
        {"$set": {"is_active": False, "end_date": end_date}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vacance non trouvée")
    return {"message": "Vacance terminée avec succès"}

# ==================== NOTIFICATIONS ROUTES ====================

@api_router.get("/notifications/settings", response_model=dict)
async def get_notification_settings(current_user: dict = Depends(get_current_user)):
    settings = await db.notification_settings.find_one(
        {"user_id": current_user['id']}, {"_id": 0}
    )
    if not settings:
        # Create default settings
        notif_settings = NotificationSettings(user_id=current_user['id'])
        doc = notif_settings.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.notification_settings.insert_one(doc)
        return doc
    return settings

@api_router.put("/notifications/settings", response_model=dict)
async def update_notification_settings(settings_data: NotificationSettingsCreate, current_user: dict = Depends(get_current_user)):
    result = await db.notification_settings.update_one(
        {"user_id": current_user['id']},
        {"$set": settings_data.model_dump()},
        upsert=True
    )
    return {"message": "Paramètres mis à jour avec succès"}

@api_router.get("/notifications", response_model=List[dict])
async def get_notifications(current_user: dict = Depends(get_current_user)):
    notifications = await db.notifications.find(
        {"user_id": current_user['id']}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return notifications

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.notifications.update_one(
        {"id": notification_id, "user_id": current_user['id']},
        {"$set": {"is_read": True}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notification non trouvée")
    return {"message": "Notification marquée comme lue"}

@api_router.put("/notifications/read-all")
async def mark_all_notifications_read(current_user: dict = Depends(get_current_user)):
    await db.notifications.update_many(
        {"user_id": current_user['id'], "is_read": False},
        {"$set": {"is_read": True}}
    )
    return {"message": "Toutes les notifications marquées comme lues"}

# ==================== DASHBOARD ROUTES ====================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    user_id = current_user['id']
    
    # Properties stats
    total_properties = await db.properties.count_documents({"user_id": user_id})
    occupied_properties = await db.properties.count_documents({"user_id": user_id, "is_occupied": True})
    
    # Tenants count
    total_tenants = await db.tenants.count_documents({"user_id": user_id})
    
    # Active leases
    active_leases = await db.leases.count_documents({"user_id": user_id, "is_active": True})
    
    # Calculate total expected monthly rent
    leases = await db.leases.find({"user_id": user_id, "is_active": True}, {"_id": 0}).to_list(1000)
    total_monthly_rent = sum(lease.get('rent_amount', 0) + lease.get('charges', 0) for lease in leases)
    
    # Current month payments
    now = datetime.now(timezone.utc)
    current_month_payments = await db.payments.find({
        "user_id": user_id,
        "period_month": now.month,
        "period_year": now.year
    }, {"_id": 0}).to_list(1000)
    total_collected = sum(p.get('amount', 0) for p in current_month_payments)
    
    # Active vacancies
    active_vacancies = await db.vacancies.count_documents({"user_id": user_id, "is_active": True})
    
    # Occupancy rate
    occupancy_rate = (occupied_properties / total_properties * 100) if total_properties > 0 else 0
    
    # Recent payments (last 6 months)
    six_months_ago = now - timedelta(days=180)
    recent_payments = await db.payments.find({
        "user_id": user_id
    }, {"_id": 0}).to_list(1000)
    
    # Group by month
    monthly_revenue = {}
    for payment in recent_payments:
        key = f"{payment['period_year']}-{payment['period_month']:02d}"
        if key not in monthly_revenue:
            monthly_revenue[key] = 0
        monthly_revenue[key] += payment.get('amount', 0)
    
    # Sort and get last 6 months
    sorted_months = sorted(monthly_revenue.items(), reverse=True)[:6]
    revenue_chart = [{"month": m[0], "amount": m[1]} for m in reversed(sorted_months)]
    
    # Unread notifications count
    unread_notifications = await db.notifications.count_documents({"user_id": user_id, "is_read": False})
    
    return {
        "total_properties": total_properties,
        "occupied_properties": occupied_properties,
        "vacant_properties": total_properties - occupied_properties,
        "total_tenants": total_tenants,
        "active_leases": active_leases,
        "total_monthly_rent": total_monthly_rent,
        "total_collected": total_collected,
        "pending_amount": total_monthly_rent - total_collected,
        "active_vacancies": active_vacancies,
        "occupancy_rate": round(occupancy_rate, 1),
        "revenue_chart": revenue_chart,
        "unread_notifications": unread_notifications
    }

# ==================== RECEIPT (QUITTANCE) GENERATION ====================

@api_router.get("/receipts/{payment_id}")
async def generate_receipt(payment_id: str, current_user: dict = Depends(get_current_user)):
    payment = await db.payments.find_one({"id": payment_id, "user_id": current_user['id']}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Paiement non trouvé")
    
    lease = await db.leases.find_one({"id": payment['lease_id']}, {"_id": 0})
    if not lease:
        raise HTTPException(status_code=404, detail="Bail non trouvé")
    
    property_doc = await db.properties.find_one({"id": lease['property_id']}, {"_id": 0})
    tenant = await db.tenants.find_one({"id": lease['tenant_id']}, {"_id": 0})
    user = await db.users.find_one({"id": current_user['id']}, {"_id": 0, "password": 0})
    
    months_fr = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", 
                 "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    
    return {
        "receipt": {
            "id": payment_id,
            "landlord_name": user['name'],
            "tenant_name": f"{tenant['first_name']} {tenant['last_name']}",
            "property_address": f"{property_doc['address']}, {property_doc['postal_code']} {property_doc['city']}",
            "property_name": property_doc['name'],
            "period": f"{months_fr[payment['period_month']]} {payment['period_year']}",
            "rent_amount": lease['rent_amount'],
            "charges": lease['charges'],
            "total_amount": payment['amount'],
            "payment_date": payment['payment_date'],
            "payment_method": payment['payment_method']
        }
    }

# ==================== EXPORT ROUTES ====================

@api_router.get("/export/payments")
async def export_payments(year: int = None, current_user: dict = Depends(get_current_user)):
    query = {"user_id": current_user['id']}
    if year:
        query["period_year"] = year
    
    payments = await db.payments.find(query, {"_id": 0}).to_list(10000)
    
    # Enrich data
    export_data = []
    for payment in payments:
        lease = await db.leases.find_one({"id": payment['lease_id']}, {"_id": 0})
        if lease:
            property_doc = await db.properties.find_one({"id": lease['property_id']}, {"_id": 0})
            tenant = await db.tenants.find_one({"id": lease['tenant_id']}, {"_id": 0})
            export_data.append({
                "date": payment['payment_date'],
                "bien": property_doc['name'] if property_doc else "",
                "locataire": f"{tenant['first_name']} {tenant['last_name']}" if tenant else "",
                "periode": f"{payment['period_month']}/{payment['period_year']}",
                "montant": payment['amount'],
                "methode": payment['payment_method']
            })
    
    return {"payments": export_data}

@api_router.get("/export/payments/excel")
async def export_payments_excel(year: int = None, current_user: dict = Depends(get_current_user)):
    """Export payments to Excel file"""
    query = {"user_id": current_user['id']}
    if year:
        query["period_year"] = year
    
    payments = await db.payments.find(query, {"_id": 0}).sort("payment_date", -1).to_list(10000)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Paiements {year if year else 'Tous'}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Date", "Bien", "Locataire", "Période", "Montant (€)", "Méthode de paiement"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    # Data rows
    total_amount = 0
    row = 2
    for payment in payments:
        lease = await db.leases.find_one({"id": payment['lease_id']}, {"_id": 0})
        if lease:
            property_doc = await db.properties.find_one({"id": lease['property_id']}, {"_id": 0})
            tenant = await db.tenants.find_one({"id": lease['tenant_id']}, {"_id": 0})
            
            months_fr = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", 
                        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            
            data = [
                payment['payment_date'],
                property_doc['name'] if property_doc else "",
                f"{tenant['first_name']} {tenant['last_name']}" if tenant else "",
                f"{months_fr[payment['period_month']]} {payment['period_year']}",
                payment['amount'],
                payment['payment_method']
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col == 5:  # Amount column
                    cell.number_format = '#,##0.00 €'
                    total_amount += value
            row += 1
    
    # Total row
    ws.cell(row=row + 1, column=4, value="TOTAL").font = Font(bold=True)
    total_cell = ws.cell(row=row + 1, column=5, value=total_amount)
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00 €'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"paiements_{year if year else 'tous'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== EMAIL REMINDER ROUTES ====================

def send_email_smtp(smtp_email: str, smtp_password: str, to_email: str, subject: str, html_content: str):
    """Send email using Gmail SMTP"""
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = smtp_email
        msg['To'] = to_email
        
        html_part = MIMEText(html_content, 'html')
        msg.attach(html_part)
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(smtp_email, smtp_password)
            server.sendmail(smtp_email, to_email, msg.as_string())
        
        return True
    except Exception as e:
        logger.error(f"Failed to send email: {e}")
        return False

@api_router.post("/reminders/test-smtp")
async def test_smtp_connection(current_user: dict = Depends(get_current_user)):
    """Test SMTP connection with saved settings"""
    settings = await db.notification_settings.find_one(
        {"user_id": current_user['id']}, {"_id": 0}
    )
    
    if not settings or not settings.get('smtp_email') or not settings.get('smtp_password'):
        raise HTTPException(status_code=400, detail="Configuration SMTP manquante")
    
    # Try to send a test email
    html_content = """
    <html>
    <body style="font-family: Arial, sans-serif; padding: 20px;">
        <h2 style="color: #064E3B;">Test de connexion RentMaestro</h2>
        <p>Si vous recevez cet email, votre configuration SMTP est correcte !</p>
        <p style="color: #78716C; font-size: 12px;">Email envoyé depuis RentMaestro</p>
    </body>
    </html>
    """
    
    success = send_email_smtp(
        settings['smtp_email'],
        settings['smtp_password'],
        settings['smtp_email'],
        "Test RentMaestro - Configuration SMTP",
        html_content
    )
    
    if success:
        # Update smtp_configured flag
        await db.notification_settings.update_one(
            {"user_id": current_user['id']},
            {"$set": {"smtp_configured": True}}
        )
        return {"message": "Email de test envoyé avec succès", "success": True}
    else:
        raise HTTPException(status_code=400, detail="Échec de l'envoi. Vérifiez vos identifiants.")

@api_router.post("/reminders/send")
async def send_payment_reminders(current_user: dict = Depends(get_current_user)):
    """Send payment reminders for unpaid rents"""
    settings = await db.notification_settings.find_one(
        {"user_id": current_user['id']}, {"_id": 0}
    )
    
    if not settings or not settings.get('smtp_configured'):
        raise HTTPException(status_code=400, detail="Configuration SMTP non validée")
    
    # Get active leases
    leases = await db.leases.find({"user_id": current_user['id'], "is_active": True}, {"_id": 0}).to_list(1000)
    
    now = datetime.now(timezone.utc)
    current_month = now.month
    current_year = now.year
    
    emails_sent = 0
    errors = []
    
    for lease in leases:
        # Check if payment exists for current month
        payment = await db.payments.find_one({
            "lease_id": lease['id'],
            "period_month": current_month,
            "period_year": current_year
        })
        
        if not payment:
            # No payment for this month - send reminder
            tenant = await db.tenants.find_one({"id": lease['tenant_id']}, {"_id": 0})
            property_doc = await db.properties.find_one({"id": lease['property_id']}, {"_id": 0})
            user = await db.users.find_one({"id": current_user['id']}, {"_id": 0, "password": 0})
            
            if tenant and tenant.get('email'):
                months_fr = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", 
                            "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
                
                total_rent = lease['rent_amount'] + lease.get('charges', 0)
                
                html_content = f"""
                <html>
                <body style="font-family: Arial, sans-serif; padding: 20px; max-width: 600px; margin: 0 auto;">
                    <div style="background: #064E3B; color: white; padding: 20px; border-radius: 8px 8px 0 0;">
                        <h1 style="margin: 0;">Rappel de loyer</h1>
                    </div>
                    <div style="border: 1px solid #E7E5E4; border-top: none; padding: 30px; border-radius: 0 0 8px 8px;">
                        <p>Bonjour {tenant['first_name']} {tenant['last_name']},</p>
                        
                        <p>Nous vous rappelons que le loyer pour le mois de <strong>{months_fr[current_month]} {current_year}</strong> 
                        n'a pas encore été enregistré pour le bien :</p>
                        
                        <div style="background: #F5F5F4; padding: 15px; border-radius: 8px; margin: 20px 0;">
                            <p style="margin: 0;"><strong>{property_doc['name']}</strong></p>
                            <p style="margin: 5px 0 0 0; color: #78716C;">{property_doc['address']}, {property_doc['postal_code']} {property_doc['city']}</p>
                        </div>
                        
                        <p><strong>Montant attendu :</strong> {total_rent:.2f} €</p>
                        <p style="font-size: 14px; color: #78716C;">
                            (Loyer : {lease['rent_amount']:.2f} € + Charges : {lease.get('charges', 0):.2f} €)
                        </p>
                        
                        <p>Merci de procéder au règlement dans les meilleurs délais.</p>
                        
                        <p>Cordialement,<br><strong>{user['name']}</strong></p>
                    </div>
                    <p style="color: #78716C; font-size: 11px; text-align: center; margin-top: 20px;">
                        Cet email a été envoyé automatiquement via RentMaestro
                    </p>
                </body>
                </html>
                """
                
                success = send_email_smtp(
                    settings['smtp_email'],
                    settings['smtp_password'],
                    tenant['email'],
                    f"Rappel de loyer - {months_fr[current_month]} {current_year}",
                    html_content
                )
                
                if success:
                    emails_sent += 1
                    # Create notification
                    notif = Notification(
                        user_id=current_user['id'],
                        type="reminder_sent",
                        title="Rappel envoyé",
                        message=f"Rappel de loyer envoyé à {tenant['first_name']} {tenant['last_name']} pour {property_doc['name']}",
                        related_id=lease['id']
                    )
                    notif_dict = notif.model_dump()
                    notif_dict['created_at'] = notif_dict['created_at'].isoformat()
                    await db.notifications.insert_one(notif_dict)
                else:
                    errors.append(f"Échec pour {tenant['email']}")
    
    return {
        "message": f"{emails_sent} rappel(s) envoyé(s)",
        "emails_sent": emails_sent,
        "errors": errors if errors else None
    }

@api_router.get("/reminders/pending")
async def get_pending_payments(current_user: dict = Depends(get_current_user)):
    """Get list of tenants with pending payments for current month"""
    leases = await db.leases.find({"user_id": current_user['id'], "is_active": True}, {"_id": 0}).to_list(1000)
    
    now = datetime.now(timezone.utc)
    current_month = now.month
    current_year = now.year
    
    pending = []
    
    for lease in leases:
        payment = await db.payments.find_one({
            "lease_id": lease['id'],
            "period_month": current_month,
            "period_year": current_year
        })
        
        if not payment:
            tenant = await db.tenants.find_one({"id": lease['tenant_id']}, {"_id": 0})
            property_doc = await db.properties.find_one({"id": lease['property_id']}, {"_id": 0})
            
            pending.append({
                "lease_id": lease['id'],
                "tenant": tenant,
                "property": property_doc,
                "amount_due": lease['rent_amount'] + lease.get('charges', 0),
                "period_month": current_month,
                "period_year": current_year
            })
    
    return {"pending": pending, "count": len(pending)}

# ==================== DOCUMENTS ROUTES ====================

@api_router.post("/documents/upload")
async def upload_document(
    file: UploadFile = File(...),
    name: str = Form(...),
    document_type: str = Form(...),
    related_type: str = Form(...),
    related_id: str = Form(...),
    notes: str = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Upload a document"""
    # Validate file size (max 10MB)
    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 10MB)")
    
    # Generate unique filename
    file_ext = Path(file.filename).suffix
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = UPLOADS_DIR / unique_filename
    
    # Save file
    with open(file_path, "wb") as f:
        f.write(contents)
    
    # Create document record
    doc = Document(
        name=name,
        document_type=document_type,
        related_type=related_type,
        related_id=related_id,
        notes=notes,
        user_id=current_user['id'],
        filename=unique_filename,
        file_size=len(contents),
        mime_type=file.content_type or "application/octet-stream"
    )
    
    doc_dict = doc.model_dump()
    doc_dict['created_at'] = doc_dict['created_at'].isoformat()
    await db.documents.insert_one(doc_dict)
    
    return {"id": doc.id, "message": "Document uploadé avec succès"}

@api_router.get("/documents")
async def get_documents(
    related_type: str = None,
    related_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all documents or filtered by related entity"""
    query = {"user_id": current_user['id']}
    if related_type:
        query["related_type"] = related_type
    if related_id:
        query["related_id"] = related_id
    
    documents = await db.documents.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return documents

@api_router.get("/documents/{document_id}")
async def get_document(document_id: str, current_user: dict = Depends(get_current_user)):
    """Get document metadata"""
    doc = await db.documents.find_one(
        {"id": document_id, "user_id": current_user['id']}, {"_id": 0}
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Document non trouvé")
    return doc

@api_router.get("/documents/{document_id}/download")
async def download_document(document_id: str, current_user: dict = Depends(get_current_user)):
    """Download a document file"""
    doc = await db.documents.find_one(
        {"id": document_id, "user_id": current_user['id']}, {"_id": 0}
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Document non trouvé")
    
    file_path = UPLOADS_DIR / doc['filename']
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Fichier non trouvé")
    
    return FileResponse(
        file_path,
        media_type=doc['mime_type'],
        filename=f"{doc['name']}{Path(doc['filename']).suffix}"
    )

@api_router.delete("/documents/{document_id}")
async def delete_document(document_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a document"""
    doc = await db.documents.find_one(
        {"id": document_id, "user_id": current_user['id']}, {"_id": 0}
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Document non trouvé")
    
    # Delete file
    file_path = UPLOADS_DIR / doc['filename']
    if file_path.exists():
        file_path.unlink()
    
    # Delete record
    await db.documents.delete_one({"id": document_id})
    
    return {"message": "Document supprimé avec succès"}

# ==================== CALENDAR ROUTES ====================

@api_router.get("/calendar/events")
async def get_calendar_events(
    month: int = None,
    year: int = None,
    current_user: dict = Depends(get_current_user)
):
    """Get calendar events for a specific month or all upcoming events"""
    user_id = current_user['id']
    events = []
    
    now = datetime.now(timezone.utc)
    target_month = month or now.month
    target_year = year or now.year
    
    # Get active leases for payment due dates
    leases = await db.leases.find({"user_id": user_id, "is_active": True}, {"_id": 0}).to_list(1000)
    
    for lease in leases:
        property_doc = await db.properties.find_one({"id": lease['property_id']}, {"_id": 0})
        tenant = await db.tenants.find_one({"id": lease['tenant_id']}, {"_id": 0})
        
        if property_doc and tenant:
            # Payment due date
            payment_day = lease.get('payment_day', 1)
            due_date = f"{target_year}-{target_month:02d}-{payment_day:02d}"
            
            # Check if payment exists
            payment = await db.payments.find_one({
                "lease_id": lease['id'],
                "period_month": target_month,
                "period_year": target_year
            })
            
            events.append({
                "id": f"payment-{lease['id']}-{target_month}-{target_year}",
                "title": f"Loyer - {property_doc['name']}",
                "date": due_date,
                "type": "payment_due" if not payment else "payment_done",
                "related_id": lease['id'],
                "property_name": property_doc['name'],
                "tenant_name": f"{tenant['first_name']} {tenant['last_name']}",
                "amount": lease['rent_amount'] + lease.get('charges', 0),
                "is_paid": payment is not None
            })
            
            # Lease end date if within next 3 months
            if lease.get('end_date'):
                end_date = datetime.fromisoformat(lease['end_date'].replace('Z', '+00:00') if isinstance(lease['end_date'], str) else lease['end_date'].isoformat())
                if end_date.month == target_month and end_date.year == target_year:
                    events.append({
                        "id": f"lease-end-{lease['id']}",
                        "title": f"Fin de bail - {property_doc['name']}",
                        "date": lease['end_date'][:10] if isinstance(lease['end_date'], str) else lease['end_date'].strftime('%Y-%m-%d'),
                        "type": "lease_end",
                        "related_id": lease['id'],
                        "property_name": property_doc['name'],
                        "tenant_name": f"{tenant['first_name']} {tenant['last_name']}",
                        "amount": None
                    })
    
    # Get active vacancies
    vacancies = await db.vacancies.find({"user_id": user_id, "is_active": True}, {"_id": 0}).to_list(1000)
    
    for vacancy in vacancies:
        property_doc = await db.properties.find_one({"id": vacancy['property_id']}, {"_id": 0})
        if property_doc:
            start_date = vacancy['start_date']
            if isinstance(start_date, str):
                start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            else:
                start_dt = start_date
            
            if start_dt.month == target_month and start_dt.year == target_year:
                events.append({
                    "id": f"vacancy-{vacancy['id']}",
                    "title": f"Vacance - {property_doc['name']}",
                    "date": start_date[:10] if isinstance(start_date, str) else start_date.strftime('%Y-%m-%d'),
                    "type": "vacancy",
                    "related_id": vacancy['id'],
                    "property_name": property_doc['name'],
                    "tenant_name": None,
                    "amount": None
                })
    
    return {"events": events, "month": target_month, "year": target_year}

# ==================== AUTOMATED REMINDERS ====================

async def send_automated_reminders():
    """Background task to send automated payment reminders"""
    logger.info("Running automated reminders check...")
    
    # Get all users with email reminders enabled
    users_settings = await db.notification_settings.find({
        "email_reminders": True,
        "smtp_configured": True
    }, {"_id": 0}).to_list(1000)
    
    for settings in users_settings:
        user_id = settings['user_id']
        frequency = settings.get('reminder_frequency', 'weekly')
        
        # Check if it's time to send based on frequency
        now = datetime.now(timezone.utc)
        
        # For weekly: send on Mondays
        # For monthly: send on 1st of month
        # For daily: always send
        should_send = False
        if frequency == 'daily':
            should_send = True
        elif frequency == 'weekly' and now.weekday() == 0:  # Monday
            should_send = True
        elif frequency == 'monthly' and now.day == 1:
            should_send = True
        
        if not should_send:
            continue
        
        # Get active leases for this user
        leases = await db.leases.find({"user_id": user_id, "is_active": True}, {"_id": 0}).to_list(1000)
        
        current_month = now.month
        current_year = now.year
        
        for lease in leases:
            # Check if payment exists for current month
            payment = await db.payments.find_one({
                "lease_id": lease['id'],
                "period_month": current_month,
                "period_year": current_year
            })
            
            if not payment:
                tenant = await db.tenants.find_one({"id": lease['tenant_id']}, {"_id": 0})
                property_doc = await db.properties.find_one({"id": lease['property_id']}, {"_id": 0})
                user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
                
                if tenant and tenant.get('email') and property_doc and user:
                    months_fr = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", 
                                "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
                    
                    total_rent = lease['rent_amount'] + lease.get('charges', 0)
                    
                    html_content = f"""
                    <html>
                    <body style="font-family: Arial, sans-serif; padding: 20px; max-width: 600px; margin: 0 auto;">
                        <div style="background: #064E3B; color: white; padding: 20px; border-radius: 8px 8px 0 0;">
                            <h1 style="margin: 0;">Rappel automatique de loyer</h1>
                        </div>
                        <div style="border: 1px solid #E7E5E4; border-top: none; padding: 30px; border-radius: 0 0 8px 8px;">
                            <p>Bonjour {tenant['first_name']} {tenant['last_name']},</p>
                            <p>Ceci est un rappel automatique concernant le loyer du mois de <strong>{months_fr[current_month]} {current_year}</strong>.</p>
                            <div style="background: #F5F5F4; padding: 15px; border-radius: 8px; margin: 20px 0;">
                                <p style="margin: 0;"><strong>{property_doc['name']}</strong></p>
                                <p style="margin: 5px 0 0 0; color: #78716C;">{property_doc['address']}</p>
                            </div>
                            <p><strong>Montant attendu :</strong> {total_rent:.2f} €</p>
                            <p>Cordialement,<br><strong>{user['name']}</strong></p>
                        </div>
                    </body>
                    </html>
                    """
                    
                    success = send_email_smtp(
                        settings['smtp_email'],
                        settings['smtp_password'],
                        tenant['email'],
                        f"[Rappel Auto] Loyer - {months_fr[current_month]} {current_year}",
                        html_content
                    )
                    
                    if success:
                        logger.info(f"Auto reminder sent to {tenant['email']}")
                    else:
                        logger.error(f"Failed to send auto reminder to {tenant['email']}")

# ==================== HEALTH CHECK ====================

@api_router.get("/")
async def root():
    return {"message": "RentMaestro API v1.0", "status": "healthy"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_event():
    """Start the scheduler for automated reminders"""
    # Run automated reminders every day at 9:00 AM
    scheduler.add_job(
        send_automated_reminders,
        CronTrigger(hour=9, minute=0),
        id="automated_reminders",
        replace_existing=True
    )
    scheduler.start()
    logger.info("Scheduler started for automated reminders")

@app.on_event("shutdown")
async def shutdown_db_client():
    scheduler.shutdown()
    client.close()
